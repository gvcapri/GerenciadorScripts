import copy
import os
import re
import tempfile
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.formula.translate import Translator, TranslatorError

QUADRA_ALIASES = {
    "Quadra",
    "QD",
    "Nº Quadra",
    "Número da Quadra",
    "Numero Quadra",
}
LOTE_ALIASES = {
    "Lote",
    "Lotes",
    "LT",
    "Nº Lote",
    "Número do Lote",
    "Numero Lote",
}
LINHAS_PARA_LOCALIZAR_CABECALHO = 30


def remover_acentos(texto: str) -> str:
    normalizado = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in normalizado if not unicodedata.combining(c))


def normalizar_cabecalho(valor: Any) -> str:
    if valor is None:
        return ""
    texto = remover_acentos(str(valor)).casefold()
    return re.sub(r"[^a-z0-9]+", "", texto)


QUADRA_NORMALIZADA = {normalizar_cabecalho(v) for v in QUADRA_ALIASES}
LOTE_NORMALIZADO = {normalizar_cabecalho(v) for v in LOTE_ALIASES}


def _romano_para_inteiro(texto: str) -> int | None:
    valores = {"I": 1, "V": 5, "X": 10, "L": 50, "C": 100, "D": 500, "M": 1000}
    if not texto or any(c not in valores for c in texto):
        return None
    total = 0
    anterior = 0
    for caractere in reversed(texto):
        atual = valores[caractere]
        if atual < anterior:
            total -= atual
        else:
            total += atual
            anterior = atual
    milhares = "M" * (total // 1000)
    resto = total % 1000
    centenas = ("", "C", "CC", "CCC", "CD", "D", "DC", "DCC", "DCCC", "CM")
    dezenas = ("", "X", "XX", "XXX", "XL", "L", "LX", "LXX", "LXXX", "XC")
    unidades = ("", "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX")
    canonico = milhares + centenas[resto // 100] + dezenas[(resto % 100) // 10] + unidades[resto % 10]
    return total if canonico == texto else None


def _converter_romano_final(texto: str) -> str:
    correspondencia = re.fullmatch(r"(.*[A-Z])([\s.\-/]+)([IVXLCDM]+)", texto)
    if not correspondencia:
        return texto
    numero = _romano_para_inteiro(correspondencia.group(3))
    if numero is None:
        return texto
    return f"{correspondencia.group(1)} {numero}"


def chave_natural(valor: Any) -> tuple[Any, ...]:
    if valor is None:
        return (9, ())
    texto_original = str(valor)
    if not texto_original.strip():
        return (9, ())

    texto = remover_acentos(texto_original).upper().strip()
    texto = _converter_romano_final(texto)
    texto = re.sub(r"(?<=\d)\s+E\s+(?=\d)", " ", texto)

    if re.match(r"^\d", texto):
        categoria = 0
    elif re.fullmatch(r"[A-Z]+", re.sub(r"[^A-Z]", "", texto)) and not re.search(r"\d", texto):
        blocos = re.findall(r"[A-Z]+", texto)
        categoria = 1 if len(blocos) == 1 else 3
    elif re.search(r"\d", texto):
        categoria = 2
    else:
        categoria = 3

    tokens: list[tuple[int, Any]] = []
    for token in re.findall(r"\d+|[A-Z]+", texto):
        if token.isdigit():
            tokens.append((0, int(token)))
        else:
            tokens.append((1, token))
    return (categoria, tuple(tokens))


def localizar_cabecalho(ws: Any) -> tuple[int, int, int] | None:
    limite = min(ws.max_row, LINHAS_PARA_LOCALIZAR_CABECALHO)
    for numero_linha in range(1, limite + 1):
        coluna_quadra = None
        coluna_lote = None
        for celula in ws[numero_linha]:
            nome = normalizar_cabecalho(celula.value)
            if coluna_quadra is None and nome in QUADRA_NORMALIZADA:
                coluna_quadra = celula.column
            if coluna_lote is None and nome in LOTE_NORMALIZADO:
                coluna_lote = celula.column
        if coluna_quadra is not None and coluna_lote is not None:
            return numero_linha, coluna_quadra, coluna_lote
    return None


def _mesclagem_no_corpo(ws: Any, primeira_linha_dados: int) -> str | None:
    for intervalo in ws.merged_cells.ranges:
        if intervalo.max_row >= primeira_linha_dados:
            return str(intervalo)
    return None


def _traduzir_formula(valor: Any, origem: str, destino: str) -> Any:
    if not (isinstance(valor, str) and valor.startswith("=")) or origem == destino:
        return valor
    try:
        return Translator(valor, origin=origem).translate_formula(destino)
    except (TranslatorError, TypeError, ValueError):
        return valor


@dataclass
class CelulaCopiada:
    valor: Any
    estilo: Any
    comentario: Any
    hiperlink: Any


@dataclass
class LinhaCopiada:
    numero_original: int
    celulas: list[CelulaCopiada]
    dimensao: Any
    chave_quadra: tuple[Any, ...]
    chave_lote: tuple[Any, ...]


@dataclass
class ResultadoAba:
    nome: str
    status: str
    detalhes: str
    linhas: int = 0


@dataclass
class ResultadoArquivo:
    origem: Path
    status: str
    destino: Path | None = None
    detalhes: str = ""
    abas: list[ResultadoAba] = field(default_factory=list)


def _capturar_linha(
    ws: Any,
    numero_linha: int,
    max_coluna: int,
    coluna_quadra: int,
    coluna_lote: int,
) -> LinhaCopiada:
    celulas = []
    for coluna in range(1, max_coluna + 1):
        celula: Cell = ws.cell(numero_linha, coluna)
        celulas.append(
            CelulaCopiada(
                valor=copy.copy(celula.value),
                estilo=copy.copy(celula._style),
                comentario=copy.copy(celula.comment),
                hiperlink=copy.copy(celula.hyperlink),
            )
        )
    return LinhaCopiada(
        numero_original=numero_linha,
        celulas=celulas,
        dimensao=copy.copy(ws.row_dimensions[numero_linha]),
        chave_quadra=chave_natural(ws.cell(numero_linha, coluna_quadra).value),
        chave_lote=chave_natural(ws.cell(numero_linha, coluna_lote).value),
    )


def _escrever_linha(ws: Any, linha: LinhaCopiada, destino: int) -> None:
    for indice, copiada in enumerate(linha.celulas, start=1):
        celula_destino: Cell = ws.cell(destino, indice)
        origem = ws.cell(linha.numero_original, indice).coordinate
        celula_destino.value = _traduzir_formula(copiada.valor, origem, celula_destino.coordinate)
        celula_destino._style = copy.copy(copiada.estilo)
        celula_destino.comment = copy.copy(copiada.comentario)
        celula_destino._hyperlink = copy.copy(copiada.hiperlink)

    dimensao = copy.copy(linha.dimensao)
    dimensao.index = destino
    ws.row_dimensions[destino] = dimensao


def _valor_assinavel(valor: Any) -> tuple[str, str]:
    return type(valor).__name__, repr(valor)


def _formula_canonica(valor: Any, coordenada: str, coluna: int) -> Any:
    if isinstance(valor, str) and valor.startswith("="):
        destino = f"{ws_coluna(coluna)}1"
        return _traduzir_formula(valor, coordenada, destino)
    return valor


def ws_coluna(numero: int) -> str:
    letras = ""
    while numero:
        numero, resto = divmod(numero - 1, 26)
        letras = chr(65 + resto) + letras
    return letras


def _assinatura_linha(ws: Any, linha: int, max_coluna: int) -> tuple[Any, ...]:
    assinatura = []
    for coluna in range(1, max_coluna + 1):
        celula = ws.cell(linha, coluna)
        valor = _formula_canonica(celula.value, celula.coordinate, coluna)
        assinatura.append(_valor_assinavel(valor))
    return tuple(assinatura)


def _validar_ordenacao(
    ws: Any,
    primeira_linha: int,
    ultima_linha: int,
    coluna_quadra: int,
    coluna_lote: int,
) -> None:
    chaves = [
        (
            chave_natural(ws.cell(linha, coluna_quadra).value),
            chave_natural(ws.cell(linha, coluna_lote).value),
        )
        for linha in range(primeira_linha, ultima_linha + 1)
    ]
    if chaves != sorted(chaves):
        raise ValueError("a conferência detectou dados fora da ordem natural")


def ordenar_aba(ws: Any, cabecalho: tuple[int, int, int]) -> ResultadoAba:
    linha_cabecalho, coluna_quadra, coluna_lote = cabecalho
    primeira = linha_cabecalho + 1
    ultima = ws.max_row
    max_coluna = ws.max_column
    if primeira > ultima:
        return ResultadoAba(ws.title, "ignorada", "não há linhas após o cabeçalho")

    mesclagem = _mesclagem_no_corpo(ws, primeira)
    if mesclagem:
        return ResultadoAba(
            ws.title,
            "ignorada",
            f"mesclagem {mesclagem} alcança o corpo dos dados",
        )

    quantidade_antes = ultima - primeira + 1
    assinaturas_antes = Counter(
        _assinatura_linha(ws, linha, max_coluna) for linha in range(primeira, ultima + 1)
    )
    pares_antes = Counter(
        (
            _valor_assinavel(ws.cell(linha, coluna_quadra).value),
            _valor_assinavel(ws.cell(linha, coluna_lote).value),
        )
        for linha in range(primeira, ultima + 1)
    )

    linhas = [
        _capturar_linha(ws, linha, max_coluna, coluna_quadra, coluna_lote)
        for linha in range(primeira, ultima + 1)
    ]
    linhas.sort(key=lambda item: (item.chave_quadra, item.chave_lote, item.numero_original))
    for destino, linha in enumerate(linhas, start=primeira):
        _escrever_linha(ws, linha, destino)

    if ws.max_row != ultima or ws.max_column != max_coluna:
        raise ValueError("a quantidade de linhas ou colunas foi alterada")
    quantidade_depois = ws.max_row - primeira + 1
    if quantidade_antes != quantidade_depois:
        raise ValueError("a quantidade de linhas de dados foi alterada")

    assinaturas_depois = Counter(
        _assinatura_linha(ws, linha, max_coluna) for linha in range(primeira, ultima + 1)
    )
    if assinaturas_antes != assinaturas_depois:
        raise ValueError("o conjunto de linhas mudou durante a ordenação")

    pares_depois = Counter(
        (
            _valor_assinavel(ws.cell(linha, coluna_quadra).value),
            _valor_assinavel(ws.cell(linha, coluna_lote).value),
        )
        for linha in range(primeira, ultima + 1)
    )
    if pares_antes != pares_depois:
        raise ValueError("o conjunto de pares quadra + lote foi alterado")
    _validar_ordenacao(ws, primeira, ultima, coluna_quadra, coluna_lote)
    return ResultadoAba(ws.title, "ordenada", "validações concluídas", quantidade_antes)


def _nome_destino(
    origem: Path,
    pasta_saida: str | os.PathLike[str] | None = None,
    nome_saida: str | None = None,
) -> Path:
    pasta = Path(pasta_saida).expanduser().resolve() if pasta_saida else origem.parent
    pasta.mkdir(parents=True, exist_ok=True)
    if nome_saida:
        nome_seguro = Path(nome_saida.strip()).name
        if not nome_seguro.casefold().endswith(".xlsx"):
            nome_seguro += ".xlsx"
        candidato = pasta / nome_seguro
    else:
        candidato = pasta / f"{origem.stem}_ORDENADO.xlsx"
    if not candidato.exists():
        return candidato
    instante = datetime.now().strftime("%Y%m%d_%H%M%S")
    base = candidato.stem
    candidato = pasta / f"{base}_{instante}.xlsx"
    contador = 2
    while candidato.exists():
        candidato = pasta / f"{base}_{instante}_{contador}.xlsx"
        contador += 1
    return candidato


def processar_arquivo(
    caminho: str | os.PathLike[str],
    pasta_saida: str | os.PathLike[str] | None = None,
    nome_saida: str | None = None,
) -> ResultadoArquivo:
    origem = Path(caminho).expanduser().resolve()
    if origem.suffix.casefold() != ".xlsx":
        return ResultadoArquivo(origem, "ignorado", detalhes="a extensão não é .xlsx")
    if not origem.is_file():
        return ResultadoArquivo(origem, "erro", detalhes="arquivo não encontrado")

    temporario: Path | None = None
    workbook = None
    try:
        try:
            workbook = load_workbook(origem, data_only=False, keep_links=True, rich_text=True)
        except TypeError:
            workbook = load_workbook(origem, data_only=False, keep_links=True)

        resultados_abas: list[ResultadoAba] = []
        for ws in workbook.worksheets:
            if normalizar_cabecalho(ws.title) == "controle":
                resultados_abas.append(ResultadoAba(ws.title, "preservada", "aba auxiliar Controle"))
                continue
            cabecalho = localizar_cabecalho(ws)
            if cabecalho is None:
                resultados_abas.append(
                    ResultadoAba(ws.title, "preservada", "colunas de quadra e lote não localizadas")
                )
                continue
            resultados_abas.append(ordenar_aba(ws, cabecalho))

        if not any(item.status == "ordenada" for item in resultados_abas):
            workbook.close()
            return ResultadoArquivo(
                origem,
                "ignorado",
                detalhes="nenhuma aba pôde ser ordenada",
                abas=resultados_abas,
            )

        destino = _nome_destino(origem, pasta_saida, nome_saida)
        descritor, nome_temp = tempfile.mkstemp(
            prefix=f".{origem.stem}_ordenando_", suffix=".xlsx", dir=destino.parent
        )
        os.close(descritor)
        temporario = Path(nome_temp)
        workbook.save(temporario)
        workbook.close()
        workbook = None

        conferencia = load_workbook(temporario, data_only=False, read_only=True, keep_links=True)
        conferencia.close()
        os.replace(temporario, destino)
        temporario = None

        conferencia_final = load_workbook(destino, data_only=False, read_only=True, keep_links=True)
        conferencia_final.close()
        return ResultadoArquivo(origem, "processado", destino, abas=resultados_abas)
    except Exception as erro:
        if workbook is not None:
            workbook.close()
        if temporario is not None and temporario.exists():
            temporario.unlink()
        return ResultadoArquivo(origem, "erro", detalhes=f"{type(erro).__name__}: {erro}")
