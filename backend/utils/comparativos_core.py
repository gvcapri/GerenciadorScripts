from __future__ import annotations

import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from docx import Document
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


M2 = "m" + chr(178)
BLUE = "233A5E"
GREEN = "91B522"
WHITE = "FFFFFF"
LIGHT_BLUE = "DCE6F1"
OK_FILL = "C6EFCE"
BAD_FILL = "FFC7CE"
WARN_FILL = "FFEB9C"
INFO_FILL = "D9EAF7"


@dataclass
class Record:
    origin: str
    quadra: str
    lote: str
    area: Decimal | None = None
    perimeter: Decimal | None = None
    front: str = ""
    street: str = ""
    latitude: Decimal | None = None
    longitude: Decimal | None = None
    latitude_places: int | None = None
    longitude_places: int | None = None
    beneficiary: str = ""
    cpf: str = ""
    process_status: str = ""
    restriction: str = ""
    source_row: int | str = ""
    extra: dict[str, object] = field(default_factory=dict)

    @property
    def key(self) -> tuple[str, str]:
        return normalize_quadra(self.quadra), normalize_lote(self.lote)


def clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("_x000D_", " ")).strip()


def deaccent(value: object) -> str:
    return "".join(
        char
        for char in unicodedata.normalize("NFKD", clean(value))
        if not unicodedata.combining(char)
    )


def column_key(value: object) -> str:
    return re.sub(r"[^A-Z0-9]+", "", deaccent(value).upper())


def normalize_quadra(value: object) -> str:
    text = deaccent(value).upper().strip()
    text = re.sub(r"\.0$", "", text)
    text = re.sub(r"\s*-\s*", "-", text)
    text = re.sub(r"\s+", " ", text)
    if re.fullmatch(r"\d+", text):
        return str(int(text))
    code = re.fullmatch(r"([A-Z]+)\s*-?\s*(\d+)", text)
    if code:
        return f"{code.group(1)}-{int(code.group(2)):02d}"
    return re.sub(r"[^A-Z0-9]+", "", text)


def readable_lot(value: object) -> str:
    text = clean(value)
    text = re.sub(r"(?i)^CHAC(?:ARA)?\.?\s*0*(\d+)", r"Chácara \1", text)
    text = re.sub(r"(?i)^AREA\s*COMUNITARIA\s*-?\s*CEMITERIO$", "Área Comunitária - Cemitério", text)
    text = re.sub(r"(?i)^AREA\s*DE\s*LAZER$", "Área de Lazer", text)
    text = re.sub(r"(?<=\d)(?=[A-ZÁÀÂÃÉÊÍÓÔÕÚÇ][a-záàâãéêíóôõúç]{2})", " ", text)
    return text


def normalize_lote(value: object) -> str:
    text = deaccent(value).upper().strip()
    text = re.sub(r"\.0$", "", text)
    text = re.sub(r"^(?:CHACARA|CHAC\.?|CH)\s*", "", text)
    text = text.replace("/", "-").replace(",", "-")
    text = re.sub(r"\s*-\s*", "-", text)
    text = re.sub(r"\s+", "", text)
    description = re.fullmatch(r"(\d+)-[A-Z ]{4,}.*", text)
    if description:
        text = description.group(1)
    aliases = {
        "AREACOMUNITARIA-CEMITERIO": "AREACOMUNITARIACEMITERIO",
        "AREACOMUNITARIACEMITERIO": "AREACOMUNITARIACEMITERIO",
        "AREADELAZER": "AREADELAZER",
    }
    if text in aliases:
        return aliases[text]
    parts = []
    for part in filter(None, text.split("-")):
        match = re.fullmatch(r"0*(\d+)([A-Z]+)?", part)
        parts.append(str(int(match.group(1))) + (match.group(2) or "") if match else part)
    return "-".join(parts)


def decimal_places(value: object) -> int | None:
    text = clean(value)
    if not text:
        return None
    decimal_part = re.search(r"[.,](\d+)\s*$", text)
    return len(decimal_part.group(1)) if decimal_part else 0


def parse_number(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = deaccent(value).replace(M2, "").replace("M2", "")
    text = re.sub(r"[^0-9,.-]", "", text)
    if not text or text in {"-", ".", ","}:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".") if text.rfind(",") > text.rfind(".") else text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def first_column(headers: list[object], aliases: Iterable[str]) -> int | None:
    normalized = [column_key(header) for header in headers]
    alias_keys = [column_key(alias) for alias in aliases]
    for alias in alias_keys:
        for index, header in enumerate(normalized):
            if header == alias:
                return index
    for alias in alias_keys:
        for index, header in enumerate(normalized):
            if alias and (alias in header or header in alias):
                return index
    return None


COLUMN_ALIASES = {
    "quadra": ["quadra", "qd", "nº quadra", "numero quadra", "num quadra"],
    "lote": ["lote", "lotes", "lt", "nº lote", "numero lote", "num lote"],
    "area": ["área", "area", "área do lote", "area do lote", "área m²", "area m2"],
    "perimeter": ["perímetro", "perimetro", "perímetro do lote", "perimetro do lote", "perímetro m"],
    "front": ["frente", "rua", "logradouro", "endereço", "endereco", "via"],
    "beneficiary": ["nome beneficiário", "nome do beneficiário", "beneficiário", "beneficiario", "nome"],
    "cpf": ["cpf beneficiário", "cpf do beneficiário", "cpf"],
    "process_status": ["status processo", "status", "situação", "situacao"],
    "latitude": ["latitude", "lat", "coord y", "coordenada y", "y"],
    "longitude": ["longitude", "long", "lon", "coord x", "coordenada x", "x"],
    "restriction": ["restrição", "restricao", "restrição ambiental", "quadro ambiental", "ambiental"],
}


def locate_sheet_and_header(path: Path):
    workbook = load_workbook(path, data_only=True, read_only=True)
    best = None
    available = list(workbook.sheetnames)
    for sheet_name in available:
        sheet = workbook[sheet_name]
        for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 30), values_only=True), 1):
            headers = list(row)
            quadra = first_column(headers, COLUMN_ALIASES["quadra"])
            lote = first_column(headers, COLUMN_ALIASES["lote"])
            score = sum(first_column(headers, aliases) is not None for aliases in COLUMN_ALIASES.values())
            if quadra is not None and lote is not None and (best is None or score > best[0]):
                best = (score, sheet_name, row_index, headers)
    if best is None:
        workbook.close()
        raise ValueError(
            "Não encontrei colunas de Quadra e Lote. Abas disponíveis: " + ", ".join(available)
        )
    return workbook, best[1], best[2], best[3]


def read_excel(path: Path, origin: str) -> list[Record]:
    workbook, sheet_name, header_row, headers = locate_sheet_and_header(path)
    sheet = workbook[sheet_name]
    columns = {name: first_column(headers, aliases) for name, aliases in COLUMN_ALIASES.items()}
    records = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
        def value(name: str):
            index = columns[name]
            return row[index] if index is not None and index < len(row) else None

        quadra, lote = clean(value("quadra")), readable_lot(value("lote"))
        if not quadra and not lote:
            continue
        records.append(
            Record(
                origin=origin,
                quadra=quadra,
                lote=lote,
                area=parse_number(value("area")),
                perimeter=parse_number(value("perimeter")),
                front=clean(value("front")),
                street=clean(value("front")),
                latitude=parse_number(value("latitude")),
                longitude=parse_number(value("longitude")),
                latitude_places=decimal_places(value("latitude")),
                longitude_places=decimal_places(value("longitude")),
                beneficiary=clean(value("beneficiary")),
                cpf=clean(value("cpf")),
                process_status=clean(value("process_status")),
                restriction=clean(value("restriction")),
                source_row=row_number,
                extra={"sheet": sheet_name},
            )
        )
    workbook.close()
    if not records:
        raise ValueError(f"Nenhum registro válido foi encontrado na planilha {path.name}.")
    return records


def docx_blocks(path: Path) -> list[str]:
    document = Document(path)
    blocks = []
    for paragraph in document.element.body.xpath(".//w:p"):
        text = clean("".join(node.text or "" for node in paragraph.xpath(".//w:t")))
        if text:
            blocks.append(text)
    return blocks


def parse_front(text: str) -> str:
    normalized = deaccent(text).lower()
    start = normalized.find("de frente")
    if start < 0:
        return ""
    segment = text[start:]
    match = re.search(
        r"confrontando\s+com\s+(.+?)(?=,?\s*(?:da[ií]\s+deflete|fechando|segue\s+confrontando|"
        r"at[ée]\s+o\s+v[ée]rtice)|[.;]|$)",
        segment,
        re.IGNORECASE,
    )
    return clean(match.group(1)).rstrip(" ,.;") if match else ""


def read_memorial(path: Path) -> list[Record]:
    paragraphs = docx_blocks(path)
    records = []
    current_text = ""
    current_q = current_l = ""

    def flush():
        nonlocal current_text, current_q, current_l
        if not (current_q or current_l):
            return
        area = re.search(r"(?:Á|A|�)REA(?:\s+TOTAL)?\s*:\s*([0-9.,]+)", current_text, re.IGNORECASE)
        perimeter = re.search(r"PER(?:Í|I|�)METRO\s*:\s*([0-9.,]+)", current_text, re.IGNORECASE)
        records.append(
            Record(
                origin="MD",
                quadra=current_q,
                lote=readable_lot(current_l),
                area=parse_number(area.group(1)) if area else None,
                perimeter=parse_number(perimeter.group(1)) if perimeter else None,
                front=parse_front(current_text),
                source_row=len(records) + 1,
            )
        )
        current_text = ""
        current_q = current_l = ""

    property_pattern = re.compile(
        r"(?:PROPRIEDADE\s*:\s*)?LOTE\s*:\s*(.*?)\s*-\s*QUADRA\s*:\s*(.*?)"
        r"(?=(?:MUNIC.PIO|BAIRRO|LOTEAMENTO|(?:Á|A|�)REA|PER.METRO)\s*:|$)",
        re.IGNORECASE,
    )
    reverse_pattern = re.compile(r"QUADRA\s*:\s*(.*?)\s*-\s*LOTE\s*:\s*(.*?)(?=(?:Á|A|�)REA\s*:|$)", re.IGNORECASE)
    for paragraph in paragraphs:
        found = property_pattern.search(paragraph) or reverse_pattern.search(paragraph)
        if found:
            flush()
            if property_pattern.search(paragraph):
                current_l, current_q = clean(found.group(1)), clean(found.group(2))
            else:
                current_q, current_l = clean(found.group(1)), clean(found.group(2))
            current_text = paragraph
        elif current_q or current_l:
            current_text += " " + paragraph
    flush()
    if not records:
        raise ValueError("Nenhum bloco LOTE/QUADRA foi encontrado no memorial de lotes.")
    return records


def read_crf(path: Path) -> list[Record]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return read_excel(path, "CRF")
    text = "\n".join(docx_blocks(path))
    starts = list(re.finditer(r"INDICA[CÇ][ÃA]O\s+NUM[ÉE]RICA\s*:?", text, re.IGNORECASE))
    chunks = [text[m.start() : starts[i + 1].start() if i + 1 < len(starts) else len(text)] for i, m in enumerate(starts)]
    if not chunks:
        chunks = [text]
    records = []
    for chunk in chunks:
        pairs = list(
            re.finditer(
                r"QUADRA\s*[:\-]?\s*([A-ZÀ-Ý0-9 ._-]+?)\s+(?:LOTE|LOTES)\s*[:\-]?\s*"
                r"([A-ZÀ-Ý0-9 /,._-]+?)(?=\s+(?:[ÁA]REA|PER[ÍI]METRO|RESTRI[CÇ][ÃA]O|QUADRA|INDICA[CÇ][ÃA]O)|\n|$)",
                chunk,
                re.IGNORECASE,
            )
        )
        area = re.search(r"[ÁA]REA(?:\s+TOTAL)?\s*:?\s*([0-9.,]+)", chunk, re.IGNORECASE)
        perimeter = re.search(r"PER[ÍI]METRO\s*:?\s*([0-9.,]+)", chunk, re.IGNORECASE)
        restriction = re.search(r"RESTRI[CÇ][ÃA]O(?:\s+AMBIENTAL)?\s*:?\s*(.+?)(?=\n|$)", chunk, re.IGNORECASE)
        for pair in pairs:
            quadra = clean(pair.group(1))
            lots = [clean(item) for item in re.split(r"\s*[,;/]\s*", pair.group(2)) if clean(item)]
            for lote in lots:
                records.append(
                    Record(
                        origin="CRF",
                        quadra=quadra,
                        lote=readable_lot(lote),
                        area=parse_number(area.group(1)) if area else None,
                        perimeter=parse_number(perimeter.group(1)) if perimeter else None,
                        restriction=clean(restriction.group(1)) if restriction else "",
                        source_row=len(records) + 1,
                    )
                )
    if not records:
        raise ValueError("Nenhuma Indicação Numérica com Quadra e Lote foi encontrada no CRF.")
    return records


def index_records(records: list[Record]) -> dict[tuple[str, str], list[Record]]:
    indexed: dict[tuple[str, str], list[Record]] = defaultdict(list)
    for record in records:
        indexed[record.key].append(record)
    return indexed


def visible_number(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def origin_article(origin: str) -> str:
    return "na" if origin in {"MT", "SHP", "CRF", "Indicação Numérica"} else "no"


def only_status(origin: str) -> str:
    if origin == "MD":
        return "Somente no MD"
    if origin == "Ecoleta":
        return "Somente no Ecoleta"
    return f"Somente {origin_article(origin)} {origin}"


def compare_value(left: Decimal | None, right: Decimal | None) -> tuple[Decimal | None, str]:
    if left is None or right is None:
        return None, "Valor ausente"
    difference = left - right
    return difference, "OK" if difference == 0 else "Divergente"


def duplicate_rows(*groups: tuple[str, list[Record]]) -> list[dict[str, object]]:
    rows = []
    for origin, records in groups:
        for key, items in index_records(records).items():
            if len(items) <= 1:
                continue
            for item in items:
                rows.append(
                    {
                        "Origem": origin,
                        f"Quadra {origin}": item.quadra,
                        f"Lote {origin}": item.lote,
                        "Quantidade na chave": len(items),
                        "Linha de origem": item.source_row,
                        "Status": "Duplicado",
                    }
                )
    return rows


def compare_sources(
    left: list[Record],
    right: list[Record],
    left_name: str,
    right_name: str,
    fields: tuple[str, ...],
    base_name: str,
) -> dict[str, list[dict[str, object]]]:
    left_index, right_index = index_records(left), index_records(right)
    base = right if base_name == right_name else left
    ordered_keys = []
    seen = set()
    for record in base:
        if record.key not in seen:
            ordered_keys.append(record.key)
            seen.add(record.key)

    comparison = []
    only_left = []
    only_right = []
    for key in ordered_keys:
        left_record = left_index.get(key, [None])[0]
        right_record = right_index.get(key, [None])[0]
        if left_record and right_record:
            row: dict[str, object] = {
                f"Quadra {left_name}": left_record.quadra,
                f"Lote {left_name}": left_record.lote,
                f"Quadra {right_name}": right_record.quadra,
                f"Lote {right_name}": right_record.lote,
            }
            statuses = []
            for field_name in fields:
                attribute, label, unit = {
                    "area": ("area", "Área", M2),
                    "perimeter": ("perimeter", "Perímetro", "m"),
                }[field_name]
                left_value = getattr(left_record, attribute)
                right_value = getattr(right_record, attribute)
                difference, status = compare_value(left_value, right_value)
                row[f"{label} {left_name} ({unit})"] = visible_number(left_value)
                row[f"{label} {right_name} ({unit})"] = visible_number(right_value)
                row[f"Dif. {label} ({unit})"] = visible_number(difference)
                statuses.append(status)
            row["Status"] = "Divergente" if "Divergente" in statuses else ("Valor ausente" if "Valor ausente" in statuses else "OK")
            comparison.append(row)
        elif left_record:
            only_left.append(
                {
                    f"Quadra {left_name}": left_record.quadra,
                    f"Lote {left_name}": left_record.lote,
                    "Status": only_status(left_name),
                }
            )
        elif right_record:
            only_right.append(
                {
                    f"Quadra {right_name}": right_record.quadra,
                    f"Lote {right_name}": right_record.lote,
                    "Status": only_status(right_name),
                }
            )

    for key, items in left_index.items():
        if key not in right_index and key not in seen:
            record = items[0]
            only_left.append({f"Quadra {left_name}": record.quadra, f"Lote {left_name}": record.lote, "Status": only_status(left_name)})
    for key, items in right_index.items():
        if key not in left_index and key not in seen:
            record = items[0]
            only_right.append({f"Quadra {right_name}": record.quadra, f"Lote {right_name}": record.lote, "Status": only_status(right_name)})

    # Revalidação obrigatória: mesmo lote e mesma área com quadra diferente é
    # divergência de quadra, não dois registros falsamente classificados como "Somente".
    reconciled_left: set[tuple[str, str]] = set()
    reconciled_right: set[tuple[str, str]] = set()
    if "area" in fields:
        left_signatures: dict[tuple[str, Decimal], list[Record]] = defaultdict(list)
        right_signatures: dict[tuple[str, Decimal], list[Record]] = defaultdict(list)
        for key, items in left_index.items():
            if key not in right_index and items[0].area is not None:
                left_signatures[(key[1], items[0].area)].append(items[0])
        for key, items in right_index.items():
            if key not in left_index and items[0].area is not None:
                right_signatures[(key[1], items[0].area)].append(items[0])
        for signature in left_signatures.keys() & right_signatures.keys():
            if len(left_signatures[signature]) != 1 or len(right_signatures[signature]) != 1:
                continue
            left_record, right_record = left_signatures[signature][0], right_signatures[signature][0]
            if left_record.key[0] == right_record.key[0]:
                continue
            row = {
                f"Quadra {left_name}": left_record.quadra,
                f"Lote {left_name}": left_record.lote,
                f"Quadra {right_name}": right_record.quadra,
                f"Lote {right_name}": right_record.lote,
            }
            for field_name in fields:
                attribute, label, unit = {
                    "area": ("area", "Área", M2),
                    "perimeter": ("perimeter", "Perímetro", "m"),
                }[field_name]
                left_value, right_value = getattr(left_record, attribute), getattr(right_record, attribute)
                difference, _ = compare_value(left_value, right_value)
                row[f"{label} {left_name} ({unit})"] = visible_number(left_value)
                row[f"{label} {right_name} ({unit})"] = visible_number(right_value)
                row[f"Dif. {label} ({unit})"] = visible_number(difference)
            row["Status"] = "Divergente: Quadra"
            comparison.append(row)
            reconciled_left.add(left_record.key)
            reconciled_right.add(right_record.key)
        only_left = [
            row for row in only_left
            if (normalize_quadra(row[f"Quadra {left_name}"]), normalize_lote(row[f"Lote {left_name}"])) not in reconciled_left
        ]
        only_right = [
            row for row in only_right
            if (normalize_quadra(row[f"Quadra {right_name}"]), normalize_lote(row[f"Lote {right_name}"])) not in reconciled_right
        ]

    duplicates = duplicate_rows((left_name, left), (right_name, right))
    summary = [
        {"Item": f"Registros {left_name}", "Valor": len(left)},
        {"Item": f"Registros {right_name}", "Valor": len(right)},
        {"Item": "Pares comparados", "Valor": len(comparison)},
        {"Item": f"Somente {left_name}", "Valor": len(only_left)},
        {"Item": f"Somente {right_name}", "Valor": len(only_right)},
        {"Item": "Linhas duplicadas", "Valor": len(duplicates)},
        {"Item": "Regra numérica", "Valor": "Tolerância zero"},
    ]
    return {
        "Resumo": summary,
        "Comparacao": comparison,
        f"Somente_{left_name}": only_left,
        f"Somente_{right_name}": only_right,
        "Duplicados": duplicates,
    }


def normalize_front(value: object) -> str:
    text = deaccent(value).upper()
    text = re.sub(r"\bAV\.?\b", "AVENIDA", text)
    text = re.sub(r"\bR\.?\b", "RUA", text)
    return re.sub(r"[^A-Z0-9]+", "", text)


def compare_fronts(md: list[Record], shp: list[Record]) -> dict[str, list[dict[str, object]]]:
    md_index = index_records(md)
    rows, divergences = [], []
    for shp_record in shp:
        md_record = md_index.get(shp_record.key, [None])[0]
        row = {
            "Quadra SHP": shp_record.quadra,
            "Lote SHP": shp_record.lote,
            "Frente MD": md_record.front if md_record else "",
            "Frente SHP": shp_record.front or shp_record.street,
        }
        if md_record is None:
            status = "Faltando em MD"
        elif not row["Frente MD"]:
            status = "Valor ausente no MD"
        elif not row["Frente SHP"]:
            status = "Faltando em SHP"
        elif normalize_front(row["Frente MD"]) == normalize_front(row["Frente SHP"]):
            status = "OK"
        else:
            status = "Divergente"
        row["Status"] = status
        rows.append(row)
        if status != "OK":
            divergences.append(row.copy())
    duplicates = duplicate_rows(("MD", md), ("SHP", shp))
    return {
        "Resumo": [
            {"Item": "Registros MD", "Valor": len(md)},
            {"Item": "Registros SHP", "Valor": len(shp)},
            {"Item": "Comparados", "Valor": len(rows)},
            {"Item": "Divergências/Faltantes", "Valor": len(divergences)},
        ],
        "Comparacao": rows,
        "Divergencias": divergences,
        "Duplicados": duplicates,
    }


def rounded_coordinate(value: Decimal | None, places: int | None) -> Decimal | None:
    if value is None or places is None:
        return None
    quantum = Decimal(1).scaleb(-places)
    return value.quantize(quantum)


def compare_shp_ecoleta(shp: list[Record], ecoleta: list[Record]) -> dict[str, list[dict[str, object]]]:
    shp_index, eco_index = index_records(shp), index_records(ecoleta)
    rows, only_eco, only_shp = [], [], []
    used_shp_rows: set[int] = set()
    coordinates_available = all(
        any(record.latitude is not None and record.longitude is not None for record in source)
        for source in (shp, ecoleta)
    )
    for eco in ecoleta:
        candidates = shp_index.get(eco.key, [])
        shp_record = None
        if coordinates_available:
            for candidate in candidates:
                if candidate.source_row in used_shp_rows:
                    continue
                if (
                    rounded_coordinate(candidate.latitude, eco.latitude_places) == eco.latitude
                    and rounded_coordinate(candidate.longitude, eco.longitude_places) == eco.longitude
                ):
                    shp_record = candidate
                    break
        if shp_record is None:
            shp_record = next(
                (candidate for candidate in candidates if candidate.source_row not in used_shp_rows),
                None,
            )
        if shp_record is None:
            only_eco.append(
                {
                    "Quadra Ecoleta": eco.quadra,
                    "Lote Ecoleta": eco.lote,
                    "Nome Beneficiário": eco.beneficiary,
                    "CPF Beneficiário": eco.cpf,
                    "Status Processo": eco.process_status,
                    "Status": "Somente no Ecoleta",
                }
            )
            continue
        used_shp_rows.add(shp_record.source_row)
        used_lat = rounded_coordinate(shp_record.latitude, eco.latitude_places)
        used_lon = rounded_coordinate(shp_record.longitude, eco.longitude_places)
        if not coordinates_available:
            coordinate_status = "Não comparado"
        elif None in (used_lat, used_lon, eco.latitude, eco.longitude):
            coordinate_status = "Valor ausente"
        elif used_lat == eco.latitude and used_lon == eco.longitude:
            coordinate_status = "OK"
        else:
            coordinate_status = "Divergente"
        status = "Ambos" if coordinate_status in {"OK", "Não comparado"} else "Divergente"
        rows.append(
            {
                "Quadra SHP": shp_record.quadra,
                "Lote SHP": shp_record.lote,
                "Rua SHP": shp_record.street or "Não possui",
                "Latitude SHP": visible_number(shp_record.latitude),
                "Longitude SHP": visible_number(shp_record.longitude),
                "Latitude SHP usada": visible_number(used_lat),
                "Longitude SHP usada": visible_number(used_lon),
                "Quadra Ecoleta": eco.quadra,
                "Lote Ecoleta": eco.lote,
                "Latitude Ecoleta": visible_number(eco.latitude),
                "Longitude Ecoleta": visible_number(eco.longitude),
                "Nome Beneficiário": eco.beneficiary,
                "CPF Beneficiário": eco.cpf,
                "Status Processo": eco.process_status,
                "Status Coordenadas": coordinate_status,
                "Status": status,
            }
        )
    for shp_record in shp:
        if shp_record.source_row not in used_shp_rows:
            only_shp.append(
                {
                    "Quadra SHP": shp_record.quadra,
                    "Lote SHP": shp_record.lote,
                    "Rua SHP": shp_record.street or "Não possui",
                    "Latitude SHP": visible_number(shp_record.latitude),
                    "Longitude SHP": visible_number(shp_record.longitude),
                    "Status": "Somente na SHP",
                }
            )
    duplicates = duplicate_rows(("SHP", shp), ("Ecoleta", ecoleta))
    duplicate_shp = [row for row in duplicates if row["Origem"] == "SHP"]
    duplicate_shp_missing = []
    for key, items in shp_index.items():
        excess = max(0, len(items) - len(eco_index.get(key, [])))
        for item in items[-excess:] if excess else []:
            duplicate_shp_missing.append(
                {
                    "Quadra SHP": item.quadra,
                    "Lote SHP": item.lote,
                    "Rua SHP": item.street or "Não possui",
                    "Latitude SHP": visible_number(item.latitude),
                    "Longitude SHP": visible_number(item.longitude),
                    "Status": "Duplicado na SHP sem ocorrência correspondente no Ecoleta",
                }
            )
    only_one_base = []
    for row in only_shp:
        only_one_base.append(
            {
                "Origem": "SHP",
                "Quadra SHP": row.get("Quadra SHP", ""),
                "Lote SHP": row.get("Lote SHP", ""),
                "Rua SHP": row.get("Rua SHP", ""),
                "Latitude SHP": row.get("Latitude SHP", ""),
                "Longitude SHP": row.get("Longitude SHP", ""),
                "Quadra Ecoleta": "",
                "Lote Ecoleta": "",
                "Nome Beneficiário": "",
                "CPF Beneficiário": "",
                "Status Processo": "",
                "Status": row.get("Status", "Somente na SHP"),
            }
        )
    for row in only_eco:
        only_one_base.append(
            {
                "Origem": "Ecoleta",
                "Quadra SHP": "",
                "Lote SHP": "",
                "Rua SHP": "",
                "Latitude SHP": "",
                "Longitude SHP": "",
                "Quadra Ecoleta": row.get("Quadra Ecoleta", ""),
                "Lote Ecoleta": row.get("Lote Ecoleta", ""),
                "Nome Beneficiário": row.get("Nome Beneficiário", ""),
                "CPF Beneficiário": row.get("CPF Beneficiário", ""),
                "Status Processo": row.get("Status Processo", ""),
                "Status": row.get("Status", "Somente no Ecoleta"),
            }
        )
    return {
        "Resumo": [
            {"Item": "Registros SHP", "Valor": len(shp)},
            {"Item": "Registros Ecoleta", "Valor": len(ecoleta)},
            {"Item": "Pares comparados", "Valor": len(rows)},
            {"Item": "Somente SHP", "Valor": len(only_shp)},
            {"Item": "Somente Ecoleta", "Valor": len(only_eco)},
            {"Item": "Coordenadas", "Valor": "Comparadas conforme casas decimais do Ecoleta" if coordinates_available else "Não comparadas: colunas ausentes em pelo menos uma base"},
        ],
        "Comparacao": rows,
        "Somente_SHP_Ecoleta": only_one_base,
        "Duplicados_SHP": duplicate_shp,
        "Duplicados_SHP_Nao_Ecoleta": duplicate_shp_missing,
        "Duplicados": duplicates,
    }


def compare_crf_shp(crf: list[Record], shp: list[Record]) -> dict[str, list[dict[str, object]]]:
    sheets = compare_sources(crf, shp, "CRF", "SHP", ("area",), "SHP")
    crf_index, shp_index = index_records(crf), index_records(shp)
    for row in sheets["Comparacao"]:
        key = (normalize_quadra(row["Quadra SHP"]), normalize_lote(row["Lote SHP"]))
        crf_record, shp_record = crf_index[key][0], shp_index[key][0]
        row["Restrição CRF"] = crf_record.restriction
        row["Rua SHP"] = shp_record.street or "Não possui"
    for row in sheets["Somente_CRF"]:
        key = (normalize_quadra(row["Quadra CRF"]), normalize_lote(row["Lote CRF"]))
        row["Restrição CRF"] = crf_index[key][0].restriction
    return sheets


def compare_crf_md(crf: list[Record], md: list[Record]) -> dict[str, list[dict[str, object]]]:
    """Compara a Indicação Numérica da CRF com o memorial, no layout aprovado."""
    sheets = compare_sources(
        crf,
        md,
        "Indicação Numérica",
        "MD",
        ("area", "perimeter"),
        "MD",
    )
    only_key = "Somente_Indicação Numérica"
    if only_key in sheets:
        sheets["Somente_Indicacao"] = sheets.pop(only_key)
    return sheets


def make_metric_sheet(mt: list[Record], md: list[Record]) -> dict[str, list[dict[str, object]]]:
    compared = compare_sources(md, mt, "MD", "MT", ("area", "perimeter"), "MT")
    md_index = index_records(md)
    metric_rows = []
    for mt_record in mt:
        md_record = md_index.get(mt_record.key, [None])[0]
        metric_rows.append(
            {
                "Quadra": mt_record.quadra,
                "Lote": mt_record.lote,
                f"Área ({M2})": visible_number(md_record.area if md_record else mt_record.area),
                "Perímetro (m)": visible_number(md_record.perimeter if md_record else mt_record.perimeter),
                "Origem dos valores": "MD" if md_record else "MT - faltando no MD",
                "Status": "OK" if md_record else "Somente na MT",
            }
        )
    compared = {"Planilha_Metrica": metric_rows, **compared}
    return compared


def safe_sheet_name(name: str, existing: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "_", name)[:31] or "Planilha"
    candidate, index = base, 2
    while candidate in existing:
        suffix = f"_{index}"
        candidate = base[: 31 - len(suffix)] + suffix
        index += 1
    existing.add(candidate)
    return candidate


def format_workbook(workbook: Workbook) -> None:
    thin = Side(style="thin", color="D9E1F2")
    for sheet in workbook.worksheets:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[1].height = 32
        for cell in sheet[1]:
            cell.fill = PatternFill("solid", fgColor=BLUE)
            cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=Side(style="medium", color=GREEN))
        headers = {cell.column: clean(cell.value) for cell in sheet[1]}
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                header = headers.get(cell.column, "")
                cell.font = Font(name="Aptos", size=10)
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(
                    horizontal="center" if any(word in header for word in ("Quadra", "Lote", "Status", "Linha", "Quantidade")) else "left",
                    vertical="center",
                    wrap_text=True,
                )
                if any(word in header for word in ("Área", "Perímetro")) and not header.startswith("Dif."):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif header.startswith("Dif."):
                    cell.number_format = "#,##0.0000"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif "Latitude" in header or "Longitude" in header:
                    cell.number_format = "0.00000000"
        status_column = next((cell.column for cell in sheet[1] if clean(cell.value).startswith("Status")), None)
        if status_column and sheet.max_row > 1:
            letter = get_column_letter(status_column)
            status_range = f"{letter}2:{letter}{sheet.max_row}"
            sheet.conditional_formatting.add(status_range, FormulaRule(formula=[f'LEFT({letter}2,2)="OK"'], fill=PatternFill("solid", fgColor=OK_FILL)))
            sheet.conditional_formatting.add(status_range, FormulaRule(formula=[f'OR(ISNUMBER(SEARCH("Divergente",{letter}2)),ISNUMBER(SEARCH("Faltando",{letter}2)))'], fill=PatternFill("solid", fgColor=BAD_FILL)))
            sheet.conditional_formatting.add(status_range, FormulaRule(formula=[f'OR(ISNUMBER(SEARCH("Somente",{letter}2)),ISNUMBER(SEARCH("ausente",{letter}2)),ISNUMBER(SEARCH("Duplicado",{letter}2)))'], fill=PatternFill("solid", fgColor=WARN_FILL)))
        for column_cells in sheet.columns:
            letter = get_column_letter(column_cells[0].column)
            maximum = max(len(clean(cell.value)) for cell in column_cells)
            header = clean(column_cells[0].value)
            cap = 55 if any(word in header for word in ("Frente", "Rua", "Restrição", "Item", "Valor")) else 28
            sheet.column_dimensions[letter].width = min(max(maximum + 2, 11), cap)
        sheet.auto_filter.ref = sheet.dimensions


def write_result(output_path: Path, sheets: dict[str, list[dict[str, object]]]) -> Path:
    output_path = output_path.expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    existing: set[str] = set()
    for requested_name, rows in sheets.items():
        if not rows and requested_name == "Duplicados":
            continue
        sheet = workbook.create_sheet(safe_sheet_name(requested_name, existing))
        visible_rows = rows or [{"Mensagem": "Nenhum registro"}]
        headers = list(visible_rows[0].keys())
        sheet.append(headers)
        for row in visible_rows:
            sheet.append([row.get(header, "") for header in headers])
    format_workbook(workbook)
    workbook.save(output_path)
    return output_path


COMPARISON_CONFIG = {
    "shp_mt": {"name": "SHP x Métrica", "inputs": (("SHP", "excel"), ("MT", "excel"))},
    "mt_md": {"name": "Métrica x Memorial de Lotes", "inputs": (("MT", "excel"), ("MD", "docx"))},
    "front_md_shp": {"name": "Frente dos lotes - Memorial x SHP", "inputs": (("MD", "docx"), ("SHP", "excel"))},
    "eco_shp": {"name": "Ecoleta x SHP", "inputs": (("Ecoleta", "excel"), ("SHP", "excel"))},
    "metric_sheet": {"name": "Planilha Métrica - Métrica x Memorial", "inputs": (("MT", "excel"), ("MD", "docx"))},
    "crf_shp": {"name": "CRF x SHP", "inputs": (("CRF", "document"), ("SHP", "excel"))},
    "crf_md": {"name": "CRF x Memorial de Lotes", "inputs": (("CRF", "document"), ("MD", "docx"))},
}


def load_input(label: str, path: Path) -> list[Record]:
    if label == "MD":
        return read_memorial(path)
    if label == "CRF":
        return read_crf(path)
    return read_excel(path, label)


def generate_comparison(kind: str, selected_paths: dict[str, Path], output_path: Path) -> dict[str, object]:
    if kind not in COMPARISON_CONFIG:
        raise ValueError(f"Tipo de comparativo desconhecido: {kind}")
    resolved_inputs = {label: Path(path).expanduser().resolve() for label, path in selected_paths.items()}
    output_path = Path(output_path).expanduser().resolve()
    for label, path in resolved_inputs.items():
        if not path.is_file():
            raise FileNotFoundError(f"Arquivo {label} não encontrado: {path}")
        if output_path == path:
            raise ValueError("A saída deve ter outro nome. Os arquivos de entrada são somente leitura.")
    data = {label: load_input(label, path) for label, path in resolved_inputs.items()}
    if kind == "shp_mt":
        sheets = compare_sources(data["SHP"], data["MT"], "SHP", "MT", ("area",), "MT")
    elif kind == "mt_md":
        sheets = compare_sources(data["MD"], data["MT"], "MD", "MT", ("area", "perimeter"), "MT")
    elif kind == "front_md_shp":
        sheets = compare_fronts(data["MD"], data["SHP"])
    elif kind == "eco_shp":
        sheets = compare_shp_ecoleta(data["SHP"], data["Ecoleta"])
    elif kind == "metric_sheet":
        sheets = make_metric_sheet(data["MT"], data["MD"])
    elif kind == "crf_shp":
        sheets = compare_crf_shp(data["CRF"], data["SHP"])
    else:
        sheets = compare_crf_md(data["CRF"], data["MD"])
    result = write_result(output_path, sheets)
    counts = {label: len(records) for label, records in data.items()}
    counts["comparados"] = len(sheets.get("Comparacao", []))
    return {"output": result, "counts": counts, "sheets": list(sheets)}
